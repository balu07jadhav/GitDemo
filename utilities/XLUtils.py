import openpyxl

def getRowCount(file,LoginData):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[LoginData]
    return (sheet.max_row)

def getColumnCount(file,LoginData):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[LoginData]
    return (sheet.max_column)

def readData(file,LoginData,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[LoginData]
    return sheet.cell(row=rownum,column=columnno).value

def writeData(file,LoginData,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[LoginData]
    sheet.cell(row=rownum,column=columnno).value = data
    workbook.save(file)

